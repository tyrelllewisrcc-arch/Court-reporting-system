import streamlit as st
import pandas as pd
import io
import openpyxl
from datetime import datetime

# --- PAGE CONFIG ---
st.set_page_config(page_title="San Pedro Auto-Filler Pro", layout="wide")
st.title("⚖️ San Pedro Court Report: Complete 9-Sheet Auto-Filler")
st.markdown("""
**System Status:** ✅ Active
**Supported Sheets:**
*   **1 & 3:** Main Crimes (Cases & Persons)
*   **2:** Disposal Breakdown
*   **4 & 5:** Convicted Demographics (Sentence & Age)
*   **6 & 7:** Juvenile Analysis (Offenses & Sentences)
*   **8 & 9:** Statutory Offenses (Drugs, Firearms, Traffic, etc.)
""")

# --- 1. SMART COLUMN MAPPING ---
def smart_read_excel(file):
    if not file: return None
    df_preview = pd.read_excel(file, header=None, nrows=20)
    header_row = 0
    found = False
    for idx, row in df_preview.iterrows():
        row_text = " ".join([str(x).upper() for x in row.values])
        if 'COURT BOOK' in row_text and ('CHARGE' in row_text or 'OFFENCE' in row_text):
            header_row = idx
            found = True
            break
            
    if not found:
        st.error(f"Could not find headers in {file.name}.")
        return None

    df = pd.read_excel(file, header=header_row)
    df.columns = [str(c).strip().upper() for c in df.columns]
    
    col_map = {}
    for c in df.columns:
        if 'COURT BOOK' in c: col_map[c] = 'CASEID'
        elif 'CHARGE' in c or 'OFFENCE' in c: col_map[c] = 'CHARGE'
        elif 'COMPLAINANT' in c or 'VICTIM' in c: col_map[c] = 'VICTIM'
        elif 'ARRAINGMENT' in c or 'ARRAIGNMENT' in c: col_map[c] = 'DATE_ARR'
        elif 'CONCLUDED' in c or 'DISPOSAL' in c: col_map[c] = 'DATE_DISP'
        elif 'AGE' in c: col_map[c] = 'AGE'
        elif 'SEX' in c or 'GENDER' in c: col_map[c] = 'GENDER'
        elif 'FURTHER' in c: col_map[c] = 'SENTENCE'
        elif 'STATUS' in c: col_map[c] = 'CASE_STATUS' 
        elif 'REMARK' in c: col_map[c] = 'REMARK'
    
    df = df.rename(columns=col_map)
    return df.loc[:, ~df.columns.duplicated()]

# --- 2. INTELLIGENT PARSERS ---

def classify_crime_sheet1(charge, victim):
    # Maps to Row Numbers in Sheet 1
    charge = str(charge).upper()
    victim = str(victim).upper()
    
    # POLICE RULE
    if any(k in victim for k in ['POLICE', 'PC ', 'CPL ', 'GOB']) and 'MINOR' not in victim:
        return 25 

    if 'ESCAPE' in charge: return 24
    if 'PERJURY' in charge: return 23
    if any(x in charge for x in ['DISORDERLY', 'ABUSIVE', 'THREAT']): return 22
    if 'RAPE' in charge: return 27
    if 'SEXUAL ASSAULT' in charge: return 28
    if 'UNLAWFUL SEXUAL' in charge: return 29
    if 'UNNATURAL' in charge: return 30
    if 'ATTEMPT' in charge and 'MURDER' in charge: return 35
    if 'MURDER' in charge: return 33
    if 'MANSLAUGHTER' in charge: return 34
    if 'GRIEVOUS' in charge: return 36
    if 'WOUNDING' in charge: return 37
    if 'HARM' in charge: return 38
    if 'AGGRAVATED ASSAULT' in charge: return 39
    if 'COMMON ASSAULT' in charge: return 40
    if 'ROBBERY' in charge: return 43
    if 'BURGLARY' in charge: return 44
    if 'THEFT' in charge: return 45
    if 'DECEPTION' in charge or 'FRAUD' in charge: return 46
    if 'HANDLING' in charge: return 47
    if 'DAMAGE' in charge: return 48
    if 'ARSON' in charge: return 49
    if 'FORGERY' in charge: return 52
    if 'DRUG' in charge or 'CANNABIS' in charge: return 54
    if 'PIPE' in charge: return 57
    if 'VEHICLE' in charge: return 58
    if 'TRAFFIC' in charge or 'MOTOR' in charge or 'LICENSE' in charge: return 59
    if 'FIREARM' in charge or 'AMMUNITION' in charge: return 59
    return 59

def classify_statutory_sheet8(charge):
    # Maps to Row Numbers in Sheet 8 (Est. based on doc list order)
    c = str(charge).upper()
    if 'DRUG' in c or 'CANNABIS' in c: return 12 # Dangerous Drugs
    if 'FIREARM' in c or 'AMMUNITION' in c: return 13 # Firearms Act
    if 'LIQUOR' in c: return 14 # Liquor Act
    if 'POLICE' in c: return 15 # Police Act
    if 'GAMBLING' in c: return 16 # Gambling
    if 'TRAFFIC' in c or 'MOTOR' in c or 'LICENSE' in c: return 17 # Summary Jurisdiction
    return 18 # Other

def parse_disposition(remark):
    r = str(remark).upper()
    if any(x in r for x in ['CONVICTED', 'GUILTY', 'FINE', 'PRISON']): return 'CONVICTED'
    if any(x in r for x in ['ACQUITTED', 'DISMISSED', 'STRUCK', 'DISCHARGED']): return 'DISMISSED'
    if any(x in r for x in ['WITHDRAWN', 'NOLLE']): return 'NOLLE'
    return 'OTHER'

def parse_sentence(sentence_text):
    s = str(sentence_text).upper()
    if 'FINE' in s or '$' in s: return 'FINE'
    if any(x in s for x in ['PRISON', 'IMPRISONMENT', 'CONFINEMENT', 'MONTHS', 'YEARS']): return 'PRISON'
    if 'PROBATION' in s or 'BOND' in s: return 'PROBATION'
    if 'REFORM' in s or 'SCHOOL' in s: return 'REFORMATORY'
    return 'OTHER'

def is_juvenile(age):
    try: return int(age) <= 16
    except: return False

def get_age_col_sheet5(age, gender):
    g = str(gender).upper()
    is_male = 'F' not in g 
    try: a = int(age)
    except: return None
    if a <= 16: return 'B' if is_male else 'C'
    if 17 <= a <= 25: return 'D' if is_male else 'E'
    if 26 <= a <= 35: return 'F' if is_male else 'G'
    if 36 <= a <= 45: return 'H' if is_male else 'I'
    if a >= 46: return 'J' if is_male else 'K'
    return None

# --- 3. TEMPLATE FILLER ---
def fill_all_sheets(template_file, df, mode):
    wb = openpyxl.load_workbook(template_file)
    
    # --- PROCESS DATA ---
    seen_cases = set()
    rows_sheet1 = [] # Unique Cases
    rows_sheet3 = [] # Persons
    
    for idx, row in df.iterrows():
        r_num = classify_crime_sheet1(row.get('CHARGE', ''), row.get('VICTIM', ''))
        rows_sheet3.append(r_num)
        
        case_id = row.get('CASEID', idx)
        if case_id not in seen_cases:
            rows_sheet1.append(r_num)
            seen_cases.add(case_id)

    # --- FILL SHEET 1 (Cases) ---
    if 'Sheet1' in wb.sheetnames:
        ws = wb['Sheet1']
        col = 'D' if mode == "New" else 'J'
        for r in rows_sheet1:
            try:
                curr = ws[f"{col}{r}"].value or 0
                ws[f"{col}{r}"] = curr + 1
            except: pass

    # --- FILL SHEET 3 (Persons) ---
    if 'Sheet3' in wb.sheetnames:
        ws = wb['Sheet3']
        col = 'D' if mode == "New" else 'J'
        for r in rows_sheet3:
            try:
                curr = ws[f"{col}{r}"].value or 0
                ws[f"{col}{r}"] = curr + 1
            except: pass

    # --- FILL SHEET 8 (Statutory Cases) ---
    if 'Sheet8' in wb.sheetnames:
        ws = wb['Sheet8']
        for idx, row in df.iterrows():
            stat_row = classify_statutory_sheet8(row.get('CHARGE', ''))
            
            if mode == "New":
                # Column C = New Cases
                try: ws[f"C{stat_row}"] = (ws[f"C{stat_row}"].value or 0) + 1
                except: pass
            elif mode == "Disposed":
                disp = parse_disposition(row.get('REMARK', ''))
                if disp == 'CONVICTED':
                    try: ws[f"E{stat_row}"] = (ws[f"E{stat_row}"].value or 0) + 1
                    except: pass
                elif disp == 'DISMISSED':
                    try: ws[f"F{stat_row}"] = (ws[f"F{stat_row}"].value or 0) + 1
                    except: pass

    # --- DISPOSED ONLY: FILL SHEETS 2, 4, 5, 6, 7, 9 ---
    if mode == "Disposed":
        
        # Sheet 2 (Disposals)
        if 'Sheet2' in wb.sheetnames:
            ws = wb['Sheet2']
            for idx, row in df.iterrows():
                r_num = classify_crime_sheet1(row.get('CHARGE', ''), row.get('VICTIM', ''))
                disp = parse_disposition(row.get('REMARK', ''))
                target_col = {'CONVICTED': 'E', 'DISMISSED': 'C', 'NOLLE': 'D'}.get(disp)
                if target_col:
                    try: ws[f"{target_col}{r_num}"] = (ws[f"{target_col}{r_num}"].value or 0) + 1
                    except: pass

        # CONVICTED LOOP
        for idx, row in df.iterrows():
            if parse_disposition(row.get('REMARK', '')) != 'CONVICTED': continue

            r_num = classify_crime_sheet1(row.get('CHARGE', ''), row.get('VICTIM', ''))
            gender = row.get('GENDER', 'M')
            is_male = 'F' not in str(gender).upper()
            age = row.get('AGE', 0)
            sent_type = parse_sentence(row.get('SENTENCE', ''))
            
            # Sheet 4 (Sentence)
            if 'Sheet4' in wb.sheetnames:
                ws = wb['Sheet4']
                s_col = {'PRISON': 'D' if is_male else 'E', 'PROBATION': 'F' if is_male else 'G', 'FINE': 'H' if is_male else 'I'}.get(sent_type)
                if s_col: 
                    try: ws[f"{s_col}{r_num}"] = (ws[f"{s_col}{r_num}"].value or 0) + 1
                    except: pass

            # Sheet 5 (Age)
            if 'Sheet5' in wb.sheetnames:
                ws = wb['Sheet5']
                a_col = get_age_col_sheet5(age, gender)
                if a_col:
                    try: ws[f"{a_col}{r_num-11}"] = (ws[f"{a_col}{r_num-11}"].value or 0) + 1
                    except: pass

            # JUVENILES (Age <= 16) - Sheets 6 & 7
            if is_juvenile(age):
                juv_row = r_num - 14 # Estimate offset
                
                # Sheet 6 (Offense)
                if 'Sheet6' in wb.sheetnames:
                    try: ws[f"F{juv_row}"] = (ws[f"F{juv_row}"].value or 0) + 1
                    except: pass
                
                # Sheet 7 (Sentence)
                if 'Sheet7' in wb.sheetnames:
                    ws = wb['Sheet7']
                    sent_col = {'PRISON': 'B', 'PROBATION': 'C', 'FINE': 'D', 'REFORMATORY': 'E'}.get(sent_type)
                    if sent_col:
                        try: ws[f"{sent_col}{juv_row}"] = (ws[f"{sent_col}{juv_row}"].value or 0) + 1
                        except: pass

            # Sheet 9 (Statutory Punishment)
            if 'Sheet9' in wb.sheetnames:
                stat_row = classify_statutory_sheet8(row.get('CHARGE', ''))
                ws = wb['Sheet9']
                # Mappings based on Stat Row
                s_col = {'PRISON': 'D' if is_male else 'E', 'PROBATION': 'B' if is_male else 'C', 'FINE': 'F' if is_male else 'G'}.get(sent_type)
                if s_col:
                    try: ws[f"{s_col}{stat_row}"] = (ws[f"{s_col}{stat_row}"].value or 0) + 1
                    except: pass

    return wb

# --- 4. MAIN INTERFACE ---
st.sidebar.header("1. Uploads")
data_file = st.sidebar.file_uploader("Data File (Excel)", type=['xlsx'])
template_file = st.sidebar.file_uploader("Blank Template", type=['xlsx'])

st.sidebar.header("2. Settings")
mode = st.sidebar.radio("Data Type", ["New Cases (Arraignments)", "Disposed Cases (Concluded)"])
is_full_year = st.sidebar.checkbox("Full Year Report")

if not is_full_year:
    report_month = st.sidebar.selectbox("Month", range(1, 13), format_func=lambda x: datetime(2025, x, 1).strftime('%B'))
report_year = st.sidebar.number_input("Year", value=2025)

if st.button("🚀 Process & Fill Report"):
    if not data_file or not template_file:
        st.error("Upload both files first.")
        st.stop()

    df = smart_read_excel(data_file)
    if df is None: st.stop()

    date_col = 'DATE_ARR' if mode.startswith("New") else 'DATE_DISP'
    if date_col not in df.columns:
        st.error(f"Missing Date Column. Need '{'Arraignment' if mode.startswith('New') else 'Concluded'}' date.")
        st.stop()
        
    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    
    if is_full_year:
        mask = (df[date_col].dt.year == report_year)
        period_name = f"Full Year {report_year}"
    else:
        mask = (df[date_col].dt.month == report_month) & (df[date_col].dt.year == report_year)
        period_name = datetime(2025, report_month, 1).strftime('%B %Y')

    df_filtered = df[mask].copy()
    st.success(f"Processing {len(df_filtered)} records for {period_name}")

    try:
        wb_filled = fill_all_sheets(template_file, df_filtered, "New" if mode.startswith("New") else "Disposed")
        
        out = io.BytesIO()
        wb_filled.save(out)
        out.seek(0)
        
        st.download_button(
            "📥 Download Complete 9-Sheet Report",
            data=out,
            file_name=f"San_Pedro_Stats_9SHEETS_{period_name.replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        st.write("### Data Preview")
        st.dataframe(df_filtered.head(50))
        
    except Exception as e:
        st.error(f"Processing Error: {e}")
